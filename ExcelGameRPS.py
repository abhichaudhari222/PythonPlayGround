from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time
from datetime import date

 
def TimeDate():
    global DateNow,TimeNow
    TimeNow= time.strftime("%H:%M:%S", time.localtime())
    # print("Current Time is :", TimeNow)
    DateNow = date.today()
    # print("Today date is: ", DateNow)

def CreateLoadWorkBook():
    

    headers       = ['Player','Selected','CPU','Result','Date','Time']
    workbook_name = 'GameExcel.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'RockPaperScissors'
    page.append(headers) # write the headers to the first line

    # Data to write:
    RockPaperScissors = [['Abhishek','Rock','Scissors','Win',DateNow,TimeNow]]

    for info in RockPaperScissors:
        page.append(info)
    wb.save(filename = workbook_name)

def AppendData():
    workbook_name = 'GameExcel.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active
    i='2'
    name="Vaishnav"
    selected="Paper"
    cpu="Rock"
    Result="Lose"
    new_RockPaperScissors = [[name,selected,cpu,Result,DateNow,TimeNow]]
    for info in new_RockPaperScissors:
        page.append(info)
    wb.save(filename=workbook_name)



TimeDate()
CreateLoadWorkBook()
# AppendData()